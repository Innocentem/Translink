from flask import (
    Blueprint, render_template, redirect, url_for, 
    flash, request, current_app, send_file
)
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from urllib.parse import urlparse, urljoin
import os
from io import BytesIO
from datetime import datetime, timedelta
import pandas as pd
from models import User, Truck, TruckRequest, ActivityLog
from forms import RegisterForm, LoginForm, TruckForm
from extensions import db
from functools import wraps
from sqlalchemy import or_
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('auth_routes.landing'))
            if current_user.role != role:
                flash('Unauthorized access!', 'danger')
                return redirect(url_for('dashboard_routes.dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

auth_routes = Blueprint('auth_routes', __name__)

@auth_routes.route('/')
@auth_routes.route('/landing', methods=['GET', 'POST'])
def landing():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_routes.analytics_dashboard'))
        return redirect(url_for('dashboard_routes.dashboard'))

    # Initialize forms
    register_form = RegisterForm()
    login_form = LoginForm()
    
    return render_template('landing.html', 
                         register_form=register_form, 
                         login_form=login_form)

# Add this helper function for URL safety check
def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc

@auth_routes.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('auth_routes.landing'))

@auth_routes.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@auth_routes.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('errors/500.html'), 500


# Create a new Blueprint for dashboard routes
dashboard_routes = Blueprint('dashboard_routes', __name__)

@dashboard_routes.route('/dashboard')
@login_required
def dashboard():
    try:
        print(f"Dashboard route - User: {current_user.username}, Role: {current_user.role}")
        
        if current_user.role == 'truck_fleet_owner':
            trucks = Truck.query.filter_by(user_id=current_user.id).all()
            truck_form = TruckForm()  # Make sure this line is present
            return render_template('dashboard.html', 
                                trucks=trucks,
                                truck_form=truck_form,  # Make sure this is being passed
                                current_user=current_user)
            
        elif current_user.role == 'transportation_service_user':
            sent_requests = TruckRequest.query.filter_by(user_id=current_user.id).all()
            return render_template('dashboard.html', 
                                sent_requests=sent_requests,
                                current_user=current_user)
            
        else:
            flash('Invalid user role!', 'danger')
            return redirect(url_for('auth_routes.landing'))
            
    except Exception as e:
        print(f"Dashboard Error: {str(e)}")
        db.session.rollback()
        flash('Error loading dashboard. Please try again.', 'danger')
        return redirect(url_for('auth_routes.landing'))

@dashboard_routes.route('/add_cargo', methods=['POST'])
@login_required
def add_cargo():
    if 'image' not in request.files:
        flash('No image file provided', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    image = request.files['image']
    if image.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    if image:
        filename = secure_filename(image.filename)
        image.save(os.path.join('static', 'uploads', filename))
        
        new_cargo = Cargo(
            name=request.form.get('name'),
            origin=request.form.get('origin'),
            destination=request.form.get('destination'),
            weight=float(request.form.get('weight')),
            image=filename,
            user_id=current_user.id,
            status='Available'
        )
        
        db.session.add(new_cargo)
        db.session.commit()
        flash('Cargo added successfully!', 'success')
    
    return redirect(url_for('dashboard_routes.dashboard'))

@dashboard_routes.route('/toggle_truck/<int:truck_id>', methods=['POST'])
@login_required
def toggle_availability(truck_id):
    truck = Truck.query.get_or_404(truck_id)
    if truck.user_id != current_user.id:
        flash('Unauthorized action!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    truck.available = not truck.available
    db.session.commit()
    flash('Truck status updated!', 'success')
    return redirect(url_for('dashboard_routes.dashboard'))

@dashboard_routes.route('/toggle_cargo/<int:cargo_id>', methods=['POST'])
@login_required
def toggle_cargo_status(cargo_id):
    cargo = Cargo.query.get_or_404(cargo_id)
    if cargo.user_id != current_user.id:
        flash('Unauthorized action!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    cargo.status = 'Available' if cargo.status == 'Transported' else 'Transported'
    db.session.commit()
    flash('Cargo status updated!', 'success')
    return redirect(url_for('dashboard_routes.dashboard'))

@dashboard_routes.route('/delete_truck/<int:truck_id>', methods=['POST'])
@login_required
def delete_truck(truck_id):
    truck = Truck.query.get_or_404(truck_id)
    if truck.user_id != current_user.id:
        flash('Unauthorized action!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    db.session.delete(truck)
    db.session.commit()
    flash('Truck deleted successfully!', 'success')
    return redirect(url_for('dashboard_routes.dashboard'))

@dashboard_routes.route('/delete_cargo/<int:cargo_id>', methods=['POST'])
@login_required
def delete_cargo(cargo_id):
    cargo = Cargo.query.get_or_404(cargo_id)
    if cargo.user_id != current_user.id:
        flash('Unauthorized action!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    db.session.delete(cargo)
    db.session.commit()
    flash('Cargo deleted successfully!', 'success')
    return redirect(url_for('dashboard_routes.dashboard'))

@dashboard_routes.route('/handle_request/<int:request_id>/<action>', methods=['POST'])
@login_required
@role_required('truck_fleet_owner')
def handle_request(request_id, action):
    try:
        truck_request = TruckRequest.query.get_or_404(request_id)
        
        # Verify ownership
        if truck_request.truck.user_id != current_user.id:
            flash('Unauthorized action!', 'danger')
            return redirect(url_for('dashboard_routes.dashboard'))
        
        if action == 'accept':
            truck_request.status = 'Accepted'
            truck_request.truck.available = False
            flash('Request accepted successfully!', 'success')
        elif action == 'reject':
            truck_request.status = 'Rejected'
            flash('Request rejected successfully!', 'success')
        else:
            flash('Invalid action!', 'danger')
            return redirect(url_for('dashboard_routes.dashboard'))
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error handling request: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_routes.dashboard'))

@dashboard_routes.route('/add_truck', methods=['POST'])
@login_required
def add_truck():
    """Add a new truck (truck fleet owners only)"""
    if current_user.role != 'truck_fleet_owner':
        flash('Unauthorized action!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    form = TruckForm()
    
    if form.validate_on_submit():
        try:
            # Check if plate number is unique
            if Truck.query.filter_by(plate_number=form.plate_number.data).first():
                flash('A truck with this plate number already exists!', 'danger')
                return redirect(url_for('dashboard_routes.dashboard'))
            
            # Save image
            image = form.image.data
            filename = secure_filename(f"{current_user.username}_{form.plate_number.data}_{image.filename}")
            image_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            image.save(image_path)
            
            # Create new truck with form data
            new_truck = Truck(
                name=form.name.data,
                plate_number=form.plate_number.data,
                driver_name=form.driver_name.data,
                driver_contact=form.driver_contact.data,  # Add this line
                routes=form.routes.data,
                image=filename,
                user_id=current_user.id,
                available=True
            )
            
            db.session.add(new_truck)
            db.session.commit()
            
            # Log activity
            ActivityLog.log_activity(
                current_user.id,
                'add_truck',
                f'Added new truck: {form.name.data} ({form.plate_number.data})'
            )
            
            flash('Truck added successfully!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding truck: {str(e)}', 'danger')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{field}: {error}', 'danger')
    
    return redirect(url_for('dashboard_routes.dashboard'))

# Create a new Blueprint for browse routes
browse_routes = Blueprint('browse_routes', __name__)

@browse_routes.route('/browse')
@login_required
def browse():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    status = request.args.get('status', '')

    query = Truck.query

    if search:
        query = query.filter(
            or_(
                Truck.name.ilike(f'%{search}%'),
                Truck.routes.ilike(f'%{search}%')
            )
        )
    
    if status:
        query = query.filter(Truck.available == (status == 'available'))

    trucks = query.paginate(
        page=page,
        per_page=9,
        error_out=False
    )

    return render_template('browse.html', 
                         trucks=trucks,
                         search=search,
                         status=status)

@browse_routes.route('/request_truck/<int:truck_id>', methods=['POST'])
@login_required
def request_truck(truck_id):
    if current_user.role != 'transportation_service_user':
        flash('Unauthorized action!', 'danger')
        return redirect(url_for('browse_routes.browse'))

    try:
        truck = Truck.query.get_or_404(truck_id)
        
        if not truck.available:
            flash('This truck is not available for booking!', 'danger')
            return redirect(url_for('browse_routes.browse'))

        # Handle cargo image upload
        cargo_image = None
        if 'cargo_image' in request.files:
            file = request.files['cargo_image']
            if file and file.filename:
                try:
                    filename = secure_filename(f"{current_user.username}_cargo_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                    file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                    cargo_image = filename
                except Exception as e:
                    print(f"Image upload error: {str(e)}")
                    # Continue without image if upload fails

        new_request = TruckRequest(
            truck_id=truck.id,
            user_id=current_user.id,
            origin=request.form['origin'],
            destination=request.form['destination'],
            cargo_details=request.form.get('cargo_details', ''),
            cargo_image=cargo_image,  # Add this field
            status='Pending'
        )

        db.session.add(new_request)
        db.session.commit()

        flash('Request submitted successfully!', 'success')
        return redirect(url_for('dashboard_routes.dashboard'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error submitting request: {str(e)}', 'danger')
        return redirect(url_for('browse_routes.browse'))

@browse_routes.route('/debug/trucks')
def debug_trucks():
    """Debug route to check trucks in database"""
    trucks = Truck.query.all()
    return {
        'total_trucks': len(trucks),
        'trucks': [{
            'id': t.id,
            'name': t.name,
            'plate_number': t.plate_number,
            'available': t.available,
            'user_id': t.user_id
        } for t in trucks]
    }

admin_routes = Blueprint('admin_routes', __name__)

@admin_routes.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    # Get date range for filtering
    days = request.args.get('days', '30')  # Default to 30 days
    end_date = datetime.now()
    start_date = end_date - timedelta(days=int(days))
    
    # Gather statistics
    stats = {
        'total_users': User.query.count(),
        'new_users': User.query.filter(User.created_at >= start_date).count(),
        'total_trucks': Truck.query.count(),
        'new_trucks': Truck.query.filter(Truck.created_at >= start_date).count(),
        'active_requests': TruckRequest.query.filter(TruckRequest.status == 'Pending').count(),
        'completed_requests': TruckRequest.query.filter(
            TruckRequest.status == 'Accepted',
            TruckRequest.request_date >= start_date
        ).count()
    }
    
    # Get recent activities
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    recent_trucks = Truck.query.order_by(Truck.created_at.desc()).limit(5).all()
    recent_requests = TruckRequest.query.order_by(TruckRequest.request_date.desc()).limit(5).all()
    
    return render_template('admin/dashboard.html',
                         stats=stats,
                         recent_users=recent_users,
                         recent_trucks=recent_trucks,
                         recent_requests=recent_requests,
                         days=days)

@admin_routes.route('/analytics')
@login_required
@role_required('admin')
def analytics_dashboard():
    try:
        # Get date range for filtering
        days = request.args.get('days', '30')
        end_date = datetime.now()
        start_date = end_date - timedelta(days=int(days))
        
        # Get all users and trucks with None checks
        users = User.query.all()
        trucks = Truck.query.all()
        
        # Get recent activities
        recent_activities = ActivityLog.query.order_by(
            ActivityLog.timestamp.desc()
        ).limit(10).all()
        
        # Calculate analytics with safe checks
        active_users = User.query.filter(
            User.last_seen.isnot(None),
            User.last_seen >= datetime.now().date()
        ).count()

        analytics = {
            'users': {
                'total': len(users),
                'new': User.query.filter(User.created_at >= start_date).count(),
                'active': active_users
            },
            'trucks': {
                'total': len(trucks),
                'available': Truck.query.filter_by(available=True).count(),
                'booked': Truck.query.filter_by(available=False).count()
            },
            'requests': {
                'total': TruckRequest.query.count() or 0,
                'accepted': TruckRequest.query.filter_by(status='Accepted').count() or 0
            }
        }
        
        # Calculate success rate safely
        total_requests = analytics['requests']['total']
        success_rate = (analytics['requests']['accepted'] / total_requests * 100) if total_requests > 0 else 0
        
        # Get truck requests
        truck_requests = TruckRequest.query.order_by(TruckRequest.request_date.desc()).all()
        
        return render_template('admin/analytics.html',
                             users=users,
                             trucks=trucks,
                             truck_requests=truck_requests,
                             analytics=analytics,
                             success_rate=success_rate,
                             days=days,
                             recent_activities=recent_activities)
                             
    except Exception as e:
        print(f"Analytics Error: {str(e)}")
        flash('Error loading analytics. Please try again.', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))

@admin_routes.route('/generate_activity_report')
@login_required
def generate_activity_report():
    if current_user.role != 'admin':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Log"

        # Add headers with styling
        headers = ['Date', 'User', 'Action', 'Details']
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Add data
        activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).all()
        row = 2
        for activity in activities:
            user = User.query.get(activity.user_id)
            ws.cell(row=row, column=1, value=activity.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=2, value=user.username if user else 'Unknown')
            ws.cell(row=row, column=3, value=activity.action)
            ws.cell(row=row, column=4, value=activity.details)
            row += 1

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'activity_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"Report generation error: {str(e)}")
        flash('Error generating activity report. Please try again.', 'danger')
        return redirect(url_for('admin_routes.analytics_dashboard'))

@admin_routes.route('/generate_metrics_report')
@login_required
def generate_metrics_report():
    if current_user.role != 'admin':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    try:
        wb = Workbook()
        
        # Users Sheet
        ws_users = wb.active
        ws_users.title = "Users"
        user_headers = ['Username', 'Role', 'Join Date', 'Last Seen']
        
        # Trucks Sheet
        ws_trucks = wb.create_sheet("Trucks")
        truck_headers = ['Name', 'Plate Number', 'Owner', 'Driver', 'Status']
        
        # Requests Sheet
        ws_requests = wb.create_sheet("Requests")
        request_headers = ['Date', 'Requester', 'Truck', 'From', 'To', 'Status']
        
        # Style headers
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Helper function to style headers
        def style_headers(worksheet, headers):
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
        
        # Add data to Users sheet
        style_headers(ws_users, user_headers)
        users = User.query.all()
        row = 2
        for user in users:
            ws_users.cell(row=row, column=1, value=user.username)
            ws_users.cell(row=row, column=2, value=user.role)
            ws_users.cell(row=row, column=3, value=user.created_at.strftime('%Y-%m-%d'))
            ws_users.cell(row=row, column=4, value=user.last_seen.strftime('%Y-%m-%d %H:%M') if user.last_seen else 'Never')
            row += 1
        
        # Add data to Trucks sheet
        style_headers(ws_trucks, truck_headers)
        trucks = Truck.query.all()
        row = 2
        for truck in trucks:
            ws_trucks.cell(row=row, column=1, value=truck.name)
            ws_trucks.cell(row=row, column=2, value=truck.plate_number)
            ws_trucks.cell(row=row, column=3, value=truck.owner.username)
            ws_trucks.cell(row=row, column=4, value=truck.driver_name)
            ws_trucks.cell(row=row, column=5, value='Available' if truck.available else 'Booked')
            row += 1
        
        # Add data to Requests sheet
        style_headers(ws_requests, request_headers)
        requests = TruckRequest.query.all()
        row = 2
        for req in requests:
            ws_requests.cell(row=row, column=1, value=req.request_date.strftime('%Y-%m-%d %H:%M'))
            ws_requests.cell(row=row, column=2, value=req.requester.username)
            ws_requests.cell(row=row, column=3, value=req.truck.name)
            ws_requests.cell(row=row, column=4, value=req.origin)
            ws_requests.cell(row=row, column=5, value=req.destination)
            ws_requests.cell(row=row, column=6, value=req.status)
            row += 1
        
        # Adjust column widths for all sheets
        for worksheet in [ws_users, ws_trucks, ws_requests]:
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'system_metrics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"Report generation error: {str(e)}")
        flash('Error generating metrics report. Please try again.', 'danger')
        return redirect(url_for('admin_routes.analytics_dashboard'))

@admin_routes.route('/manage_account', methods=['POST'])  # Remove <int:user_id>
@login_required
@role_required('admin')
def manage_account():
    if not current_user.role == 'admin':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    action = request.form.get('action')
    user_id = request.form.get('user_id')
    
    if not user_id:
        flash('No user specified!', 'danger')
        return redirect(url_for('admin_routes.analytics_dashboard'))
    
    try:
        user = User.query.get_or_404(user_id)
        
        if user.role == 'admin':
            flash('Cannot modify admin accounts!', 'danger')
            return redirect(url_for('admin_routes.analytics_dashboard'))
        
        if action == 'delete':
            # Log the deletion
            ActivityLog.log_activity(
                current_user.id,
                'delete_account',
                f'Admin deleted account: {user.username}'
            )
            db.session.delete(user)
            flash(f'Account {user.username} has been deleted.', 'success')
            
        elif action == 'suspend':
            days = int(request.form.get('suspension_days', 1))
            reason = request.form.get('suspension_reason', 'No reason provided')
            
            user.is_suspended = True
            user.suspension_end = datetime.utcnow() + timedelta(days=days)
            user.suspension_reason = reason
            
            # Log the suspension
            ActivityLog.log_activity(
                current_user.id,
                'suspend_account',
                f'Admin suspended account: {user.username} for {days} days. Reason: {reason}'
            )
            flash(f'Account {user.username} has been suspended for {days} days.', 'success')
            
        elif action == 'unsuspend':
            user.is_suspended = False
            user.suspension_end = None
            user.suspension_reason = None
            
            # Log the unsuspension
            ActivityLog.log_activity(
                current_user.id,
                'unsuspend_account',
                f'Admin unsuspended account: {user.username}'
            )
            flash(f'Account {user.username} has been unsuspended.', 'success')
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error managing account: {str(e)}', 'danger')
    
    return redirect(url_for('admin_routes.analytics_dashboard'))

@admin_routes.route('/delete_post', methods=['POST'])
@login_required
@role_required('admin')
def delete_post():
    if not current_user.role == 'admin':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('dashboard_routes.dashboard'))
    
    post_type = request.form.get('post_type')
    post_id = request.form.get('post_id')
    reason = request.form.get('deletion_reason')
    
    try:
        if post_type == 'truck':
            post = Truck.query.get_or_404(post_id)
            owner = post.owner
            post_title = post.name
        elif post_type == 'request':
            post = TruckRequest.query.get_or_404(post_id)
            owner = post.requester
            post_title = f"Request #{post.id}"
        else:
            flash('Invalid post type!', 'danger')
            return redirect(url_for('admin_routes.analytics_dashboard'))
        
        # Log the deletion
        ActivityLog.log_activity(
            current_user.id,
            f'delete_{post_type}',
            f'Admin deleted {post_type}: {post_title}. Reason: {reason}'
        )
        
        # Notify the owner (you can implement this later)
        # send_notification(owner.id, f'Your {post_type} has been removed by admin. Reason: {reason}')
        
        db.session.delete(post)
        db.session.commit()
        
        flash(f'{post_type.title()} deleted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting {post_type}: {str(e)}', 'danger')
    
    return redirect(url_for('admin_routes.analytics_dashboard'))

@admin_routes.before_request
def update_last_seen():
    if current_user.is_authenticated:
        current_user.last_seen = datetime.utcnow()
        try:
            db.session.commit()
        except:
            db.session.rollback()